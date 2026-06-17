"""
App.py — Consolidador Fiscal B3 [v2 — Threading + Custódia Inicial]
================================
Interface Streamlit para processamento de notas SINAC (Clear CTVM).
Day Trade · Mini-Índice WIN · Cálculo automático de IR / IRRF / DARF.

Melhorias v2:
• ProcessPoolExecutor → ThreadPoolExecutor (compatibilidade Streamlit)
• Input de Custódia Inicial via st.data_editor
• Processamento paralelo de PDFs (multi-core seguro)
"""

import io
import math
import unicodedata
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
import streamlit as st

from extrator_nota_corretagem import extrair_dados_nota_direto
from consolidador_notas import consolidar_notas

# ════════════════════════════════════════════════════════════════════════
#  CONFIGURAÇÃO DA PÁGINA
# ════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Consolidador Fiscal B3",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    .block-container { padding-top: 2rem; }
    div[data-testid="metric-container"] {
        background: #1e1e2e;
        border: 1px solid #2e2e3e;
        border-radius: 10px;
        padding: 1rem 1.2rem;
    }
    div[data-testid="metric-container"] label { font-size: 0.78rem; color: #888; }
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════
#  CABEÇALHO
# ════════════════════════════════════════════════════════════════════════

st.title("📊 Consolidador Fiscal B3")
st.caption(
    "Day Trade · Mini-Índice WIN · Clear CTVM — "
    "apura IR, IRRF e DARF automaticamente a partir das notas SINAC"
)
st.divider()


# ════════════════════════════════════════════════════════════════════════
#  HELPERS DE EXTRAÇÃO
# ════════════════════════════════════════════════════════════════════════

def _sem_acento(s: str) -> str:
    """Remove acentos para comparação robusta independente de encoding."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    ).upper()


def _eh_nota_sinac(texto: str) -> bool:
    """
    Verifica se o texto contém uma nota de negociação SINAC.
    Usa comparação sem acento para suportar variações de encoding do pdfplumber.
    """
    t = _sem_acento(texto)
    return "NOTA" in t and "NEGOCI" in t


def _processar_pdf(bytes_pdf: bytes) -> tuple[list[dict], list[str]]:
    """
    Extrai notas de corretagem de um PDF.

    Estratégia: tenta pdfplumber (layout=False → layout=True).
    Se falhar, usa pypdf como fallback.
    Retorna (notas_extraidas, avisos_de_debug).
    """
    notas: list[dict] = []
    avisos: list[str] = []

    # ── Tentativa 1: pdfplumber ───────────────────────────────────────────
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(bytes_pdf)) as pdf:
            for num_pag, pagina in enumerate(pdf.pages, start=1):

                # Extrai texto em dois modos; usa o que encontrar NOTA
                texto = ""
                for layout in [False, True]:
                    try:
                        t = pagina.extract_text(layout=layout) or ""
                    except Exception:
                        t = pagina.extract_text() or ""
                    if _eh_nota_sinac(t):
                        texto = t
                        break  # encontrou — não precisa tentar o outro modo

                if not texto:
                    continue  # página sem nota SINAC → pula

                d = extrair_dados_nota_direto(texto)
                if d.get("data_pregao") and d.get("operacoes"):
                    notas.append(d)
                elif d.get("data_pregao"):
                    # Encontrou data mas não operações → ajuste = 0 naquele dia
                    avisos.append(
                        f"Página {num_pag}: data {d['data_pregao']} encontrada "
                        f"mas ajuste day trade = 0. Pregão zerado ou nota sem operações."
                    )

        if notas:
            return notas, avisos

    except ImportError:
        avisos.append("pdfplumber não instalado — usando pypdf como fallback.")
    except Exception as e:
        avisos.append(f"pdfplumber falhou ({e}) — tentando pypdf.")

    # ── Fallback: pypdf ───────────────────────────────────────────────────
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(bytes_pdf))
        for num_pag, pagina in enumerate(reader.pages, start=1):
            texto = pagina.extract_text() or ""
            if not _eh_nota_sinac(texto):
                continue
            d = extrair_dados_nota_direto(texto)
            if d.get("data_pregao") and d.get("operacoes"):
                notas.append(d)
    except Exception as e:
        avisos.append(f"pypdf também falhou: {e}")

    return notas, avisos


# ════════════════════════════════════════════════════════════════════════
#  GERADOR DE PLANILHA EXCEL
# ════════════════════════════════════════════════════════════════════════

def _gerar_excel(ops: pd.DataFrame, rm: pd.DataFrame) -> bytes:
    """
    Gera planilha Excel com 2 abas:
      • Resumo Mensal    — apuração por período com IR e DARF
      • Detalhamento     — linha a linha por pregão, com formatação condicional

    Padrão financeiro: fonte Arial, largura de colunas fixas, formato R$, 
    linha de totais com fórmulas SUM, congelamento de cabeçalho.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Estilos ──────────────────────────────────────────────────────────
    H_FILL  = PatternFill("solid", start_color="1F3864")
    H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TOT_FILL = PatternFill("solid", start_color="D6E4F0")
    TOT_FONT = Font(name="Arial", bold=True, size=10)
    D_FONT   = Font(name="Arial", size=10)
    FMT_BRL  = 'R$\\ #,##0.00;[Red]\\-R$\\ #,##0.00'
    FMT_PCT  = "0%"
    _side    = Side(style="thin", color="CCCCCC")
    _border  = Border(left=_side, right=_side, top=_side, bottom=_side)
    C, R, L  = (Alignment(horizontal=x, vertical="center") for x in ("center","right","left"))

    def _hcell(ws, row, col, value, width):
        c = ws.cell(row=row, column=col, value=value)
        c.font, c.fill, c.alignment = H_FONT, H_FILL, C
        ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def _dcell(ws, row, col, value, fmt=None, align=R, fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = D_FONT
        c.alignment = align
        c.border = _border
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill
        return c

    def _tcell(ws, row, col, formula, fmt=FMT_BRL):
        c = ws.cell(row=row, column=col, value=formula)
        c.font, c.fill, c.alignment, c.border = TOT_FONT, TOT_FILL, R, _border
        c.number_format = fmt
        return c

    # ── Aba 1: Resumo Mensal ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo Mensal"
    ws1.freeze_panes = "A2"

    cols1 = [
        ("Período",              14),
        ("Pregões",              10),
        ("Resultado Líquido (R$)", 22),
        ("Base Cálculo IR (R$)", 22),
        ("Alíquota IR",          13),
        ("IRRF Retido (R$)",     18),
        ("IR a Pagar (R$)",      18),
    ]
    for i, (h, w) in enumerate(cols1, 1):
        _hcell(ws1, 1, i, h, w)

    data_row = 2
    for _, r in rm.iterrows():
        _dcell(ws1, data_row, 1, str(r["mes"]),             align=C)
        _dcell(ws1, data_row, 2, int(r["qtd_ops"]),          fmt="#,##0", align=C)
        _dcell(ws1, data_row, 3, float(r["resultado_bruto"]),fmt=FMT_BRL)
        _dcell(ws1, data_row, 4, float(r["imposto_base"]),   fmt=FMT_BRL)
        _dcell(ws1, data_row, 5, float(r["aliquota_ir"]),    fmt=FMT_PCT, align=C)
        _dcell(ws1, data_row, 6, float(r["irrf_fonte"]),     fmt=FMT_BRL)
        _dcell(ws1, data_row, 7, float(r["ir_devido"]),      fmt=FMT_BRL)
        data_row += 1

    # Linha de totais com fórmulas Excel
    tr = data_row
    _dcell(ws1, tr, 1, "TOTAL", align=C).font = TOT_FONT
    ws1.cell(row=tr, column=1).fill = TOT_FILL
    for col in [2, 3, 4, 6, 7]:
        cl = get_column_letter(col)
        _tcell(ws1, tr, col, f"=SUM({cl}2:{cl}{tr-1})",
               fmt="#,##0" if col == 2 else FMT_BRL)

    # ── Aba 2: Detalhamento por Pregão ───────────────────────────────────────
    ws2 = wb.create_sheet("Detalhamento")
    ws2.freeze_panes = "A2"

    cols2 = [
        ("Data Pregão",          14),
        ("Status",               12),
        ("Ajuste Bruto (R$)",    20),
        ("Custos B3 (R$)",       16),
        ("IRRF Retido (R$)",     18),
        ("Resultado Líq. (R$)",  20),
    ]
    for i, (h, w) in enumerate(cols2, 1):
        _hcell(ws2, 2, i, h, w)

    # Linha 1: título da aba
    ws2.merge_cells("A1:F1")
    t = ws2.cell(row=1, column=1,
                 value="Detalhamento por Pregão — Day Trade Mini-Índice WIN")
    t.font = Font(name="Arial", bold=True, size=11, color="1F3864")
    t.alignment = C

    data_row2 = 3
    ops_s = ops.sort_values("data_pregao").reset_index(drop=True)
    for _, r in ops_s.iterrows():
        res  = float(r["resultado_bruto"])
        taxa = float(r["taxa_rateada"])
        irrf = math.floor(res * 0.01 * 100) / 100 if res > 0 else 0.0
        liq  = round(res - irrf, 2)
        bruto = round(res + taxa, 2)

        fill = (PatternFill("solid", start_color="E8F5E9") if res > 0
                else PatternFill("solid", start_color="FFEBEE") if res < 0
                else PatternFill("solid", start_color="F5F5F5"))

        _dcell(ws2, data_row2, 1,
               r["data_pregao"].strftime("%d/%m/%Y"), align=C, fill=fill)
        status = "Lucro ▲" if res > 0 else ("Prejuízo ▼" if res < 0 else "Zero")
        _dcell(ws2, data_row2, 2, status, align=C, fill=fill)
        for col, val in [(3, bruto), (4, taxa), (5, irrf), (6, liq)]:
            _dcell(ws2, data_row2, col, val, fmt=FMT_BRL, fill=fill)
        data_row2 += 1

    # Totais aba 2
    tr2 = data_row2
    _dcell(ws2, tr2, 1, "TOTAL", align=C).font = TOT_FONT
    ws2.cell(row=tr2, column=1).fill = TOT_FILL
    for col in [3, 4, 5, 6]:
        cl = get_column_letter(col)
        _tcell(ws2, tr2, col, f"=SUM({cl}3:{cl}{tr2-1})")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════
#  UPLOAD
# ════════════════════════════════════════════════════════════════════════

st.subheader("1 · Envio das Notas de Corretagem")

arquivos = st.file_uploader(
    "Arraste ou selecione os PDFs das notas (Clear / SINAC)",
    type=["pdf"],
    accept_multiple_files=True,
    help=(
        "Envie todos os PDFs do mês para obter a apuração completa. "
        "Cada arquivo pode conter uma ou mais páginas."
    ),
)

if not arquivos:
    st.info("📂 Faça o upload dos PDFs para iniciar a análise fiscal.")
    st.stop()


# ════════════════════════════════════════════════════════════════════════
#  CUSTÓDIA INICIAL (NOVO)
# ════════════════════════════════════════════════════════════════════════

st.subheader("2 · Custódia Inicial (Opcional)")
st.caption("Insira o saldo e o Preço Médio de meses anteriores para Swing Trade.")

# Tabela interativa para o usuário preencher o PM anterior
df_custodia_inicial = pd.DataFrame(columns=["ticker", "tipo_ativo", "quantidade", "preco_medio"])
custodia_input = st.data_editor(
    df_custodia_inicial,
    num_rows="dynamic",
    width=None,
    column_config={
        "ticker": st.column_config.TextColumn("Ticker (Ex: PETR4)", required=False),
        "tipo_ativo": st.column_config.SelectboxColumn("Tipo", options=["ACAO", "FUTURO"], default="ACAO"),
        "quantidade": st.column_config.NumberColumn("Quantidade", min_value=1, step=1),
        "preco_medio": st.column_config.NumberColumn("Preço Médio (R$)", format="R$ %.4f", min_value=0.01),
    },
)

st.subheader("3 · Processamento em Lote")

# ════════════════════════════════════════════════════════════════════════
#  PROCESSAMENTO COM THREADS
# ════════════════════════════════════════════════════════════════════════

notas_extraidas: list[dict] = []
todos_avisos:    list[str]  = []

# Preparar inputs para as threads (nome_arquivo, bytes)
carga_trabalho = [(arq.name, arq.read()) for arq in arquivos]

def _processar_arquivo_thread(dados):
    """Wrapper para ThreadPoolExecutor."""
    nome, bts = dados
    return _processar_pdf(bts)

with st.spinner(f"Processando {len(arquivos)} PDFs em paralelo..."):
    with ThreadPoolExecutor(max_workers=4) as executor:
        resultados = list(executor.map(_processar_arquivo_thread, carga_trabalho))
        
        for idx, (notas, avisos) in enumerate(resultados):
            notas_extraidas.extend(notas)
            todos_avisos.extend([f"**{carga_trabalho[idx][0]}**: {a}" for a in avisos])

# Exibe avisos de extração (não são erros fatais)
if todos_avisos:
    with st.expander("ℹ️ Avisos de extração", expanded=False):
        for av in todos_avisos:
            st.info(av)

validas = [n for n in notas_extraidas if n.get("data_pregao") and n.get("operacoes")]

if not validas:
    st.error(
        "**Nenhuma nota de negociação foi extraída.**  \n"
        "Possíveis causas:  \n"
        "- O PDF não é uma nota SINAC da Clear CTVM  \n"
        "- A nota não possui operações de day trade registradas  \n"
        "- O PDF está protegido por senha ou corrompido"
    )
    if todos_avisos:
        st.warning("Veja os avisos acima para mais detalhes.")
    st.stop()

st.success(f"🎉  **{len(validas)} pregão(ões)** consolidado(s) com sucesso!")
st.divider()


# ════════════════════════════════════════════════════════════════════════
#  CONSOLIDAÇÃO
# ════════════════════════════════════════════════════════════════════════

try:
    # Converter input de custódia para o formato do livro interno
    livro_inicial = {}
    for _, row in custodia_input.dropna(how='all').iterrows():
        if pd.notna(row['ticker']):
            livro_inicial[row['ticker'].upper()] = {
                'qty': int(row['quantidade']),
                'avg': float(row['preco_medio']),
                'tipo_ativo': row['tipo_ativo'],
                'ultima_compra': pd.NaT
            }
    
    # Repassa a custódia inicial para o motor fiscal
    resultado = consolidar_notas(validas, livro_inicial if livro_inicial else None)
except Exception as exc:
    st.error(f"Erro na consolidação: {exc}")
    st.stop()

ops = resultado["operacoes"]
rm  = resultado["resultado_mensal"]


# ════════════════════════════════════════════════════════════════════════
#  KPIs
# ════════════════════════════════════════════════════════════════════════

st.subheader("4 · Resumo do Período")

resultado_liquido = rm["resultado_bruto"].sum()
irrf_acumulado    = rm["irrf_fonte"].sum()
ir_devido_total   = rm["ir_devido"].sum()
n_lucro = int((ops["resultado_bruto"] > 0).sum())
n_perda = int((ops["resultado_bruto"] < 0).sum())

c1, c2, c3, c4 = st.columns(4)

c1.metric(
    "Pregões analisados",
    len(validas),
    help=f"{n_lucro} com lucro · {n_perda} com perda",
)
c2.metric(
    "Resultado líquido",
    f"R$ {resultado_liquido:,.2f}",
    delta="lucro" if resultado_liquido > 0 else "prejuízo",
    delta_color="normal" if resultado_liquido > 0 else "inverse",
    help="Soma dos ajustes diários líquidos de custos B3",
)
c3.metric(
    "IRRF retido (acumulado)",
    f"R$ {irrf_acumulado:,.2f}",
    help="1% retido pela corretora em cada pregão com resultado positivo",
)
c4.metric(
    "DARF a pagar",
    f"R$ {ir_devido_total:,.2f}",
    delta="nenhum ✓" if ir_devido_total == 0 else "atenção",
    delta_color="normal" if ir_devido_total == 0 else "inverse",
    help="Imposto de Renda líquido (20% Day Trade − IRRF retido)",
)


# ════════════════════════════════════════════════════════════════════════
#  VEREDICTO FISCAL
# ════════════════════════════════════════════════════════════════════════

st.divider()
st.subheader("5 · Veredicto Fiscal")

if ir_devido_total == 0 and resultado_liquido <= 0:
    st.success(
        f"✅  **Sem DARF.**  Resultado negativo de **R$ {abs(resultado_liquido):,.2f}** no período.  \n"
        f"Escriture este prejuízo na ficha *Renda Variável → Day Trade* da DIRPF "
        f"para abater lucros tributáveis em meses futuros."
    )
    if irrf_acumulado > 0:
        st.info(
            f"ℹ️  O IRRF de **R$ {irrf_acumulado:,.2f}** já retido pela corretora nos pregões "
            f"com lucro pode ser compensado em DARFs futuros ou restituído na Declaração Anual."
        )

elif ir_devido_total == 0 and resultado_liquido > 0:
    st.success(
        f"✅  **Sem DARF.**  O IRRF retido (R$ {irrf_acumulado:,.2f}) "
        f"já cobre o imposto apurado sobre o lucro de R$ {resultado_liquido:,.2f}."
    )

else:
    st.error(
        f"⚠️  **DARF de R$ {ir_devido_total:,.2f}** a recolher.  \n"
        f"- Código DARF: **6015** — Renda Variável · Ganho Líquido  \n"
        f"- Vencimento: último dia útil do mês seguinte ao período de apuração  \n"
        f"- Base de cálculo: R$ {rm['imposto_base'].sum():,.2f}  ·  "
        f"Alíquota: 20%  ·  IRRF já retido: R$ {irrf_acumulado:,.2f}"
    )


# ════════════════════════════════════════════════════════════════════════
#  APURAÇÃO MENSAL
# ════════════════════════════════════════════════════════════════════════

st.divider()
st.subheader("6 · Apuração por Mês")

rm_show = rm[[
    "mes", "qtd_ops", "resultado_bruto",
    "imposto_base", "aliquota_ir", "irrf_fonte", "ir_devido",
]].copy()

rm_show.columns = [
    "Período", "Pregões",
    "Resultado Líq. (R$)", "Base Cálculo IR (R$)",
    "Alíquota", "IRRF Retido (R$)", "IR a Pagar (R$)",
]
rm_show["Alíquota"] = rm_show["Alíquota"].map(lambda x: f"{x*100:.0f}%")
rm_show["Período"]  = rm_show["Período"].astype(str)

st.dataframe(
    rm_show,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Resultado Líq. (R$)":   st.column_config.NumberColumn(format="R$ %.2f"),
        "Base Cálculo IR (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
        "IRRF Retido (R$)":     st.column_config.NumberColumn(format="R$ %.2f"),
        "IR a Pagar (R$)":      st.column_config.NumberColumn(format="R$ %.2f"),
    },
)


# ════════════════════════════════════════════════════════════════════════
#  DETALHAMENTO POR PREGÃO
# ════════════════════════════════════════════════════════════════════════

st.divider()
st.subheader("7 · Detalhamento por Pregão")

ops_dia = ops[["data_pregao", "taxa_rateada", "resultado_bruto"]].copy()

ops_dia["Ajuste Bruto (R$)"]   = (ops_dia["resultado_bruto"] + ops_dia["taxa_rateada"]).round(2)
ops_dia["Custos B3 (R$)"]      = ops_dia["taxa_rateada"].round(2)
ops_dia["IRRF Retido (R$)"]    = ops_dia["resultado_bruto"].apply(
    lambda r: round(r * 0.01, 2) if r > 0 else 0.0
)
ops_dia["Resultado Líq. (R$)"] = (
    ops_dia["resultado_bruto"] - ops_dia["IRRF Retido (R$)"]
).round(2)
ops_dia["Data Pregão"] = ops_dia["data_pregao"].dt.strftime("%d/%m/%Y")
ops_dia["Status"] = ops_dia["resultado_bruto"].apply(
    lambda r: "🟢 Lucro" if r > 0 else ("🔴 Prejuízo" if r < 0 else "⚪ Zero")
)

tabela = ops_dia[[
    "Data Pregão", "Status",
    "Ajuste Bruto (R$)", "Custos B3 (R$)",
    "IRRF Retido (R$)", "Resultado Líq. (R$)",
]].sort_values("Data Pregão").reset_index(drop=True)

st.dataframe(
    tabela,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Ajuste Bruto (R$)":   st.column_config.NumberColumn(format="R$ %.2f"),
        "Custos B3 (R$)":      st.column_config.NumberColumn(format="R$ %.2f"),
        "IRRF Retido (R$)":    st.column_config.NumberColumn(format="R$ %.2f"),
        "Resultado Líq. (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
    },
)


# ════════════════════════════════════════════════════════════════════════
#  EXPORTAR PLANILHA
# ════════════════════════════════════════════════════════════════════════

st.divider()
st.subheader("8 · Exportar Planilha")

try:
    excel_bytes = _gerar_excel(ops, rm)
    st.download_button(
        label="⬇️  Baixar planilha (.xlsx)",
        data=excel_bytes,
        file_name="consolidador_fiscal_b3.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Planilha com duas abas: Resumo Mensal e Detalhamento por Pregão",
    )
    st.caption(
        "A planilha contém fórmulas SUM nativas do Excel. "
        "Abra no Excel ou Google Sheets para ver os totais calculados automaticamente."
    )
except Exception as exc:
    st.warning(f"Não foi possível gerar a planilha: {exc}")

# ════════════════════════════════════════════════════════════════════════
#  RODAPÉ
# ════════════════════════════════════════════════════════════════════════

st.divider()
st.caption(
    "Cálculos com base nas notas SINAC Clear CTVM · "
    "Day Trade Mini-Índice (WIN M26) · "
    "IR 20% · IRRF 1% na fonte (IN RFB nº 1585/2015) · "
    "Isenção R$ 20k aplicável apenas a ações Swing Trade"
)
